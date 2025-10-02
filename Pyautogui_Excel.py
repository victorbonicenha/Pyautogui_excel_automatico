import os
import pyautogui
import webbrowser
import pyperclip
from openpyxl import load_workbook
from time import sleep
import requests
from dotenv import load_dotenv

# -------- CARREGAR VARIÁVEIS DO .ENV --------
load_dotenv()

CAMINHO_PLANILHA = os.getenv("PLANILHA_PATH")
ITOKEN = os.getenv("TELEGRAM_TOKEN")
CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

# -------- FUNÇÃO TELEGRAM --------
def enviar_telegram(mensagem: str):
    try:
        url = f"https://api.telegram.org/bot{ITOKEN}/sendMessage"
        payload = {"chat_id": CHAT_ID, "text": mensagem}
        requests.post(url, data=payload)
    except Exception as e:
        print("Erro ao enviar mensagem no Telegram:", e)

# -------- PLANILHA --------
wb = load_workbook(CAMINHO_PLANILHA)
ws = wb.worksheets[0]  # primeira aba
novo_nome = CAMINHO_PLANILHA  

# -------- ABRIR SITE --------
URL = "https://www8.receita.fazenda.gov.br/simplesnacional/aplicacoes.aspx?id=21"
webbrowser.open(URL)
print("Abrindo o site... maximize a janela se necessário.")
sleep(2)

primeira_vez = True  # controla a navegação
contador = 0         # conta quantos CNPJs já foram processados

# -------- LOOP --------
for row in ws.iter_rows(min_row=2, max_col=1):
    cnpj_cell = row[0]
    cnpj = str(cnpj_cell.value).strip() if cnpj_cell.value else ""

    if not cnpj:
        print(f"Linha {cnpj_cell.row} ignorada (sem CNPJ).")
        continue

    simples_col = ws[f"B{cnpj_cell.row}"].value 
    mei_col = ws[f"C{cnpj_cell.row}"].value
    if simples_col and mei_col:
        print(f"Linha {cnpj_cell.row} já preenchida -> {cnpj}")
        continue

    # -------- IR ATÉ CAMPO DE CNPJ --------
    if primeira_vez:
        for _ in range(7):  # primeira vez: 7 TAB
            pyautogui.press("tab")
            sleep(0.5)
        primeira_vez = False
    else:
        pyautogui.press("tab")  # nas próximas: só 1 TAB
        sleep(0.5)

    # -------- DIGITAR CNPJ --------
    pyautogui.hotkey("ctrl", "a")
    pyautogui.press("backspace")
    pyautogui.write(cnpj, interval=0.05)
    sleep(1)

    # -------- CONFIRMAR (TAB + ENTER) --------
    pyautogui.press("tab")
    sleep(1)
    pyautogui.press("enter")
    sleep(5)

    # -------- COPIAR RESULTADO --------
    pyautogui.hotkey("ctrl", "a")
    pyautogui.hotkey("ctrl", "c")
    sleep(1)
    texto = pyperclip.paste().lower()

    simples = "Indefinido"
    mei = "Indefinido"

    for linha in texto.splitlines():
        if "situação no simples nacional" in linha:
            if "não" in linha or "nao" in linha:
                simples = "Não"
            elif "sim" in linha:
                simples = "Sim"

        if "situação no simei" in linha:
            if "não" in linha or "nao" in linha:
                mei = "Não"
            elif "sim" in linha:
                mei = "Sim"

    ws[f"B{cnpj_cell.row}"] = simples
    ws[f"C{cnpj_cell.row}"] = mei

    contador += 1
    if contador % 50 == 0:  # salva a cada 50 linhas processadas
        wb.save(novo_nome)
        msg = f"{contador} linhas processadas e salvas ✅ (Última linha: {cnpj_cell.row})"
        print(msg)
        enviar_telegram(msg)

    print(f"Linha {cnpj_cell.row} salva -> {cnpj} | Simples={simples}, MEI={mei}")

    # -------- VOLTAR P/ MENU --------
    pyautogui.press("tab", presses=2, interval=0.2)
    pyautogui.press("enter")
    sleep(1.5)

# -------- SALVA NO FINAL --------
wb.save(novo_nome)
final_msg = f"Processo finalizado com sucesso 🚀 Total de {contador} CNPJs processados."
print(final_msg)
enviar_telegram(final_msg)
